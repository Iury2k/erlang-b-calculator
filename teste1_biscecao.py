import math
from openpyxl import Workbook

def erlang_b(A, N):
    numerator = (A**N) / math.factorial(N)
    denominator = sum([(A**k) / math.factorial(k) for k in range(N + 1)])
    return numerator / denominator

def bissection_method(equation, target, lower_bound, upper_bound, max_iterations=1000, tolerance=1e-6):
    for _ in range(max_iterations):
        midpoint = (lower_bound + upper_bound) / 2
        if abs(equation(midpoint) - target) < tolerance:
            return midpoint
        if equation(midpoint) < target:
            lower_bound = midpoint
        else:
            upper_bound = midpoint
    print("O método não convergiu após", max_iterations, "iterações.")
    return None

def create_erlang_table(max_A, max_N):
    wb = Workbook()
    ws = wb.active
    ws.title = "Erlang Table"
    
    # Escreve os cabeçalhos
    ws['A1'] = "A \\ N"
    for n in range(1, max_N + 1):
        ws.cell(row=1, column=n + 1, value=f"N={n}")
    
    # Preenche a tabela
    for a in range(1, max_A + 1):
        ws.cell(row=a + 1, column=1, value=f"A={a}")
        for n in range(1, max_N + 1):
            B = erlang_b(a, n)
            ws.cell(row=a + 1, column=n + 1, value=B)
    
    return wb

def save_excel_file(workbook, filename):
    workbook.save(filename)
    print(f"Arquivo '{filename}' salvo com sucesso!")

# Exemplo de uso
while True:
    print("\nCalculadora de Erlang B")
    print("1. Calcular N")
    print("2. Calcular A")
    print("3. Calcular B")
    print("4. Gerar e exportar tabela Erlang")
    print("5. Sair")
    option = int(input("Escolha uma opção (1/2/3/4/5): "))

    if option == 1:
        B = float(input("Digite o valor de B: "))
        A = float(input("Digite o valor de A: "))
        lower_bound = 1
        upper_bound = 1000  # Valor arbitrário, pode ser ajustado conforme necessário
        N = bissection_method(lambda x: erlang_b(A, x), B, lower_bound, upper_bound)
        if N is not None:
            print("O valor de N é:", int(N))
    elif option == 2:
        B = float(input("Digite o valor de B: "))
        N = int(input("Digite o valor de N: "))
        lower_bound = 0
        upper_bound = 1000  # Valor arbitrário, pode ser ajustado conforme necessário
        A = bissection_method(lambda x: erlang_b(x, N), B, lower_bound, upper_bound)
        if A is not None:
            print("O valor de A é:", A)
    elif option == 3:
        A = float(input("Digite o valor de A: "))
        N = int(input("Digite o valor de N: "))
        print("O valor de B é:", erlang_b(A, N))
    elif option == 4:
        max_A = int(input("Digite o valor máximo de A para a tabela: "))
        max_N = int(input("Digite o valor máximo de N para a tabela: "))
        filename = input("Digite o nome do arquivo de saída: ")
        table = create_erlang_table(max_A, max_N)
        save_excel_file(table, filename)
    elif option == 5:
        print("Encerrando o programa.")
        break
    else:
        print("Opção inválida.")
